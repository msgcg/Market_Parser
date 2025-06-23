import json
import logging
import time
import random
from bs4 import BeautifulSoup
import datetime
try:
    from .parsing import init_webdriver, save_cookies, load_cookies
    from .telegram_debugger import send_debug
except:
    from parsing import init_webdriver, save_cookies, load_cookies
    from telegram_debugger import send_debug
# Настройка логирования
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")


COOKIES_FILE = "yandex_cookies.json"


def normalize_url(url):
    """Нормализует URL, удаляя ненужные параметры запроса."""
    '''parsed_url = urlparse(url)
    query_params = parse_qs(parsed_url.query)

    # Оставляем только 'sku' и  часть пути, содержащую 'product'
    relevant_keys = ['sku']
    filtered_query_params = {key: query_params[key] for key in relevant_keys if key in query_params}

    # Собираем нормализованный URL
    normalized_url = parsed_url._replace(
        query=urlencode(filtered_query_params, doseq=True),  # doseq=True для списков
        fragment=""  # Убираем фрагмент (якорь)
    ).geturl()

    #Дополнительно удаляем параметры из самого пути
    path_parts = parsed_url.path.split('/')
    product_part = next((part for part in path_parts if part.startswith("product")), None)
    if product_part:
        normalized_url = "https://market.yandex.ru/" + product_part
        if 'sku' in filtered_query_params:
            normalized_url += "?" + urlencode(filtered_query_params, doseq=True)'''

    return url.split('?')[0]


def parse_yandex_market_category(category_url):
    """Парсит категорию Яндекс.Маркета: сначала собирает все ссылки, потом парсит."""

    logging.info(f"Starting parser for category: {category_url}")
    driver = init_webdriver()
    products = []
    normalized_links_to_process = []  # Список НОРМАЛИЗОВАННЫХ ссылок для обработки
    processed_links = set()  # Множество уже обработанных НОРМАЛИЗОВАННЫХ ссылок
    scroll_increment = 500
    scroll_position = 0

    try:
        driver.get(category_url)
        load_cookies(driver, COOKIES_FILE)
        driver.refresh()
        time.sleep(random.uniform(0.7, 1.2))
        logging.info("Main category page loaded.")

        # --- Этап 1: Прокрутка и сбор ссылок ---
        while True:
            driver.execute_script(f"window.scrollTo(0, {scroll_position});")
            scroll_position += scroll_increment
            time.sleep(random.uniform(0.3, 1))

            page_source = driver.page_source
            soup = BeautifulSoup(page_source, "html.parser")

            product_links = soup.find_all("div", {"data-cs-name": "navigate"})
            logging.info(f"Found {len(product_links)} product links on the page (scroll position: {scroll_position}).")

            for link_div in product_links:
                a_tag = link_div.find('a')
                if a_tag and 'href' in a_tag.attrs:
                    original_link = "https://market.yandex.ru" + a_tag['href']
                    normalized_link = normalize_url(original_link)

                    if normalized_link not in processed_links:
                        processed_links.add(normalized_link)
                        normalized_links_to_process.append(
                            (original_link, normalized_link))  # Сохраняем обе ссылки
                        logging.debug(f"Added to processing queue: {original_link} (normalized: {normalized_link})")


            new_height = driver.execute_script("return document.body.scrollHeight")
            if scroll_position >= new_height:
                logging.info("Reached the end of the page, stopping scroll.")
                break

        logging.info(f"Collected {len(normalized_links_to_process)} links to process.")

        # --- Этап 2: Парсинг собранных ссылок ---
        for original_link, normalized_link in normalized_links_to_process:
            logging.info(f"Processing link: {original_link}")
            try:
                driver.get(original_link)  # Переходим по оригинальной
                time.sleep(random.uniform(0.7, 1.2))
                product_page_source = driver.page_source
                product_soup = BeautifulSoup(product_page_source, "html.parser")

                # ---  Код парсинга страницы товара (как и раньше) ---
                offer_div = product_soup.find("div", {"data-zone-name": "cpa-offer", "class": "cia-vs cia-cs"})

                if offer_div and 'data-zone-data' in offer_div.attrs:
                    try:
                        data_json = json.loads(offer_div['data-zone-data'])

                        if all(key in data_json for key in ['priceDetails', 'modelId', 'marketSku']):
                            price_details = data_json.get('priceDetails', {})
                            price = price_details.get('price', {}).get('value', 'Цена не указана')
                            discounted_price = price_details.get("discountedPrice", {}).get("price", {}).get(
                                "value", None)
                            if discounted_price is not None:
                                price = discounted_price
                            old_price = price_details.get("discountedPrice", {}).get("discount", {}).get("value",
                                                                                                        "Старая цена не указана")

                            if 'Бренд' in product_soup.get_text():
                                # Если слово есть, ищем нужный элемент
                                # Находим все элементы для бренда
                                target_divs = product_soup.find(id="product-description").find_all('div', {
                                    "class": "ds-text ds-text_weight_reg ds-text_typography_text ds-text_text_loose ds-text_text_reg"})

                                # Берем второй div и извлекаем текст из span внутри него
                                second_target_div = target_divs[1] if len(target_divs) > 1 else None
                                brand = second_target_div.find('span').text if second_target_div else None
                            else:
                                brand = None

                            product = {
                                'name': product_soup.find("h1", {"data-auto": "productCardTitle"}).text.strip() if product_soup.find(
                                    "h1", {"data-auto": "productCardTitle"}) else "Название не указано",
                                'price': price,
                                'old_price': old_price,
                                'url': normalized_link,  # Сохраняем НОРМАЛИЗОВАННЫЙ URL
                                'brand': brand if brand else "Бренд не указан",
                                'description': product_soup.find("div", {"id": "product-description"}).find("div", {
                                    "class": "ds-text ds-text_weight_reg ds-text_typography_text xt_vL ds-text_text_loose ds-text_text_reg"}).text.strip() if product_soup.find(
                                    "div", {"id": "product-description"}).find("div", {
                                    "class": "ds-text ds-text_weight_reg ds-text_typography_text xt_vL ds-text_text_loose ds-text_text_reg"}) else "Описание отсутствует",
                                'modelId': data_json.get('modelId'),
                                'marketSku': data_json.get('marketSku'),
                                'productId': data_json.get('productId'),
                                'wareId': data_json.get('wareId'),
                                'shopId': data_json.get('shopId'),
                                "supplierId": data_json.get("supplierId"),
                            }

                            # Находим изображение с id="transition-page"
                            img = product_soup.find('img', {"id": "transition-page"})

                            # Проверяем, существует ли изображение и атрибут 'src'
                            if img and img.get('src'):
                                first_img_src = img['src']
                                product['imageUrl'] = first_img_src
                            else:
                                product['imageUrl'] = "Нет изображения"

                            product['update'] = f"{datetime.now().strftime("%Y-%m-%d")}"

                            # print(product['imageUrl'])

                            products.append(product)
                            logging.info(f"Product parsed successfully: {product.get('name', 'Без имени')}")

                        elif all(key in data_json for key in ['price', 'oldPrice', 'productId']):

                            if 'Бренд' in product_soup.get_text():
                                # Если слово есть, ищем нужный элемент
                                # Находим все элементы для бренда
                                target_divs = product_soup.find(id="product-description").find_all('div', {
                                    "class": "ds-text ds-text_weight_reg ds-text_typography_text ds-text_text_loose ds-text_text_reg"})

                                # Берем второй div и извлекаем текст из span внутри него
                                second_target_div = target_divs[1] if len(target_divs) > 1 else None
                                brand = second_target_div.find('span').text if second_target_div else None
                            else:
                                brand = None

                            product = {
                                'name': product_soup.find("h1", {"data-auto": "productCardTitle"}).text.strip() if product_soup.find(
                                    "h1", {"data-auto": "productCardTitle"}) else "Название не указано",
                                'price': data_json.get('price', 'Цена не указана'),
                                'old_price': data_json.get('oldPrice', 'Старая цена не указана'),
                                'url': normalized_link,  # Сохраняем НОРМАЛИЗОВАННЫЙ URL
                                'brand': brand if brand else "Бренд не указан",
                                'description': product_soup.find("div", {"id": "product-description"}).find("div", {
                                    "class": "ds-text ds-text_weight_reg ds-text_typography_text xt_vL ds-text_text_loose ds-text_text_reg"}).text.strip() if product_soup.find(
                                    "div", {"id": "product-description"}).find("div", {
                                    "class": "ds-text ds-text_weight_reg ds-text_typography_text xt_vL ds-text_text_loose ds-text_text_reg"}) else "Описание отсутствует",
                                'productId': data_json.get('productId'),
                                'wareId': data_json.get('wareId'),
                                'shopId': data_json.get('shopId'),
                                "supplierId": data_json.get("supplierId"),
                                'modelId': data_json.get('modelId'),
                                'marketSku': data_json.get('marketSku'),
                            }

                            # Находим изображение с id="transition-page"
                            img = product_soup.find('img', {"id": "transition-page"})

                            # Проверяем, существует ли изображение и атрибут 'src'
                            if img and img.get('src'):
                                first_img_src = img['src']
                                product['imageUrl'] = first_img_src
                            else:
                                product['imageUrl'] = "Нет изображения"

                            product['update'] = f"{datetime.now().strftime("%Y-%m-%d")}"

                            # print(product['imageUrl'])

                            products.append(product)
                            logging.info(f"Product parsed successfully: {product.get('name', 'Без имени')}")

                        else:
                            logging.warning(f"Неожиданная структура data-zone-data: {data_json}")
                            continue

                    except (KeyError, json.JSONDecodeError, TypeError) as e:
                        logging.error(
                            f"Error extracting data from JSON: {e}, data: {offer_div.get('data-zone-data')}")
                    except Exception as e:
                        logging.error(f"Error during product parsing: {e}")

            except Exception as e:
                logging.error(f"Error processing product page ({original_link}): {e}")
                send_debug(f"Ошибка парсинга яндекс маркета ({original_link}): {e}")

    except Exception as e:
        logging.error(f"An error occurred: {e}")
        send_debug(f"Ошибка парсинга яндекс маркета: {e}")
        return json.dumps({"error": str(e)}, ensure_ascii=False)
    finally:
        save_cookies(driver, COOKIES_FILE)
        driver.quit()

    #CATEGORY_FILE_PATH = f'parsed_data/{category_url.replace('/', '_').replace(':', '_').replace('-','_')}_products.json'
    #with open(CATEGORY_FILE_PATH, 'w', encoding='utf-8') as f:
    #    json.dump(products, f, ensure_ascii=False, indent=4)
    CATEGORY_EXCEL_FILE_PATH = f'parsed_data/yandex/{category_url[:100].replace('/', '_').replace(':', '_')}_products.xlsx'
    if len(products)!=0:
        print(f"При парсинге яндекс маркета {category_url} получено {len(products)} карточек")
        send_debug(f"При парсинге яндекс маркета {category_url} получено {len(products)} карточек")
    else:
        print(f"Ошибка парсинга яндекс маркета {category_url}")
        send_debug(f"Ошибка парсинга яндекс маркета {category_url}")
        return None
    update_excel_file(CATEGORY_EXCEL_FILE_PATH, products)

    return json.dumps(products[:100], ensure_ascii=False, indent=4) # Возвращаем первые 100 записей во избежание перенаполнения страницы


import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime


def update_excel_file(file_path, data_list):
    try:
        # Проверяем, существует ли файл, если нет, создаем новый
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Products"

        # Заголовки (если нужно)
        headers = ["Название", "Цена со скидкой", "Обычная цена", "Бренд", "Описание", "Product ID", "Ware ID", 
                   "Shop ID", "Supplier ID", "Model ID", "Market SKU", "Ссылка", "Ссылка на изображение", "Дата обновления"]
        
        if sheet.max_row == 1:  # Если первый ряд пустой, записываем заголовки
            for col_num, header in enumerate(headers, start=1):
                sheet[get_column_letter(col_num) + "1"] = header

        existing_data = {normalize_url(row[11].value.strip()): (index + 1, row[12].value) for index, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=13), start=2)}
        
        for item in data_list:
            row = [
                item["name"], item["price"], item["old_price"], item["brand"], item["description"],
                item["productId"], item["wareId"], item["shopId"], item["supplierId"], item["modelId"], item["marketSku"],
                item["url"], item["imageUrl"], item["update"]
            ]
            
            new_url = normalize_url(item["url"].strip())  # Очищаем URL от параметров
            new_date_str = item["update"]

            if not new_url:
                print(f"Пропущено: {item['name']} (отсутствует URL)")
                continue

            # Флаг, указывающий, что обновление или добавление необходимо
            update_needed = False
            add_needed = True
            row_index_to_update = None

            for existing_url, (row_index, existing_date_str) in existing_data.items():
                if existing_url == new_url:  # Нашли совпадение по основному URL
                    # Сравниваем дату обновления
                    if new_date_str != existing_date_str:
                        update_needed = True
                        row_index_to_update = row_index
                    else:
                        print(f"Пропущено: {item['name']} (актуальная дата совпадает)")
                    break  # Если URL найден, больше не проверяем

            for existing_url, (row_index, existing_date_str) in existing_data.items():
                if existing_url == new_url:  # Нашли совпадение по основному URL
                    add_needed = False
                    break  # Если URL найден, больше не проверяем

            # Если данные отличаются или строки нет в таблице — обновляем или добавляем строку
            if update_needed:
                for col_num, value in enumerate(row, start=1):
                    sheet[get_column_letter(col_num) + str(row_index_to_update)] = value
                print(f"Обновлено в Excel: {item['name']} (новая дата: {new_date_str})")
            elif add_needed:
                sheet.append(row)
                print(f"Добавлено в Excel: {item['name']}")

        workbook.save(file_path)  # Сохраняем файл
        print(f"Обновление в {file_path} завершено!")
        send_debug(f"Обновление яндекс категорий в {file_path} завершено!")

    except Exception as e:
        print(f"Ошибка при обновлении {file_path}: {e}")
        send_debug(f"Ошибка при обновлении яндекс категорий {file_path}: {e}")


