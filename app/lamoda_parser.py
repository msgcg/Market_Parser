import requests
from bs4 import BeautifulSoup
from dataclasses import dataclass, field
from typing import List, Optional
from urllib.parse import urlparse, urlunparse, urljoin,parse_qs,urlencode
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter
import os  # Import the os module
try:
    from .telegram_debugger import send_debug
except: from telegram_debugger import send_debug
try:
    from .parsing import save_cookies, load_cookies, init_webdriver
except: from parsing import save_cookies, load_cookies, init_webdriver
CATEGORY_EXCEL_FILE_PATH = ''
COOKIES_FILE = "lamoda_cookies.json"  # Путь к файлу cookies

@dataclass
class Prod:
    article: str
    brand: str
    title: str
    price: Optional[float]
    price_old: Optional[float]
    price_new: Optional[float]
    full_url: str
    update_date: str = field(default_factory=lambda: datetime.now().strftime("%Y-%m-%d"))

    def __post_init__(self):
        if self.price_new is not None and self.price_old is None:
            self.price_old = 0.0
        if self.price_old is not None and self.price_new is None:
            self.price = self.price_old
        if self.price is None:
            if self.price_new is not None:
                self.price = self.price_new
            elif self.price_old is not None:
                self.price = self.price_old


def log(message: str):
    print(f"[LOG] {message}")


def parse_lamoda_category(category_url: str) -> List[Prod]:
    global CATEGORY_EXCEL_FILE_PATH
    CATEGORY_EXCEL_FILE_PATH = f'parsed_data/lamoda/{category_url[:100].replace('/', '_').replace(':', '_')}_products.xlsx'
    """Parses a Lamoda category and extracts product information."""
    products = []
    page_number = 1
    first_product_full_url = None

    parsed_url = urlparse(category_url)
    base_url = urlunparse(parsed_url._replace(query=''))
    params = parse_qs(parsed_url.query)

    # Удаляем параметр "page", если он есть
    params.pop("page", None)

    # Кодируем параметры обратно в строку запроса
    source_params = urlencode(params, doseq=True)

    while True:
        log(f'Processing page {page_number}...')
        page_url = f"{base_url}?{source_params}&page={page_number}"

        try:
            driver = init_webdriver()
            load_cookies(driver,COOKIES_FILE)
            r = driver.get(page_url)
            save_cookies(driver,COOKIES_FILE)
        except Exception as e:
            log(f"Failed to fetch page {page_number}: {e}")
            send_debug(f"Ошибка загрузки lamoda страницы {page_number}: {e}")
            return []

        soup = BeautifulSoup(r.text, 'html.parser')
        current_page_products = __parse_page(soup, page_url)

        if not current_page_products:
            log(f"No products found on page {page_number}. Assuming end.")
            break

        current_first_product_full_url = current_page_products[0].full_url

        if page_number == 1:
            first_product_full_url = current_first_product_full_url
            products.extend(current_page_products)
            page_number += 1
            continue

        if current_first_product_full_url == first_product_full_url:
            log("First product URL matches page 1. Assuming end.")
            break
        else:
            products.extend(current_page_products)
            page_number += 1

    log(f"Total products parsed: {len(products)}")
    return products


def __parse_page(soup: BeautifulSoup, base_url: str) -> List[Prod]:
    """Parses a single page of Lamoda product listings."""
    products = []
    items = soup.select('.x-product-card__card')

    for item in items:
        try:
            link_element = item.select_one('a.x-product-card__link')
            if not link_element:
                continue
            relative_link = link_element['href']
            full_url = urljoin(base_url, relative_link)
            article = relative_link.split('/')[-2] if relative_link else "Unknown"

            brand_element = item.select_one('.x-product-card-description__brand-name')
            brand = brand_element.text.strip() if brand_element else "Unknown"

            title_element = item.select_one('.x-product-card-description__product-name')
            title = title_element.text.strip() if title_element else "Unknown"

            price_new_element, price_single_element, price_old_element = (
                item.select_one('.x-product-card-description__price-new'),
                item.select_one('.x-product-card-description__price-single'),
                item.select_one('.x-product-card-description__price-old')
            )

            price = price_old = price_new = None

            if price_new_element:
                price_new_text = price_new_element.text.replace('₽', '').replace(' ', '').strip()
                try:
                    price_new = float(price_new_text)
                except ValueError:
                    log(f"Error parsing new price '{price_new_text}'")
                    continue

                if price_old_element:
                    try:
                        price_old = float(price_old_element.text.replace('₽', '').replace(' ', '').strip())
                    except ValueError:
                        log(f"Error parsing old price")

            elif price_single_element:
                try:
                    price = float(price_single_element.text.replace('₽', '').replace(' ', '').strip())
                except ValueError:
                    log(f"Error parsing single price")
                    continue

            log(f"Parsed: {article}, {brand}, {title}, Price: {price}, Old: {price_old}, New: {price_new}, {full_url}")
            product = Prod(article, brand, title, price, price_old, price_new, full_url)
            products.append(product)

        except Exception as e:
            log(f"Error parsing a product: {e}")
            continue
    if len(products)!=0:
        print(f"Успешно спарсено {len(products)} товаров с lamoda")
        send_debug(f"Успешно спарсено {len(products)} товаров с lamoda")
    else:
        print("Error. Нет товаров Lamoda")
        send_debug("Error. Нет товаров Lamoda")
        return None
    update_lamoda_excel_file(CATEGORY_EXCEL_FILE_PATH,products)
    return products[:100] # Возвращаем первые 100 записей во избежание перенаполнения страницы


def normalize_url(url: str) -> str:
    return url.split('?')[0]


def update_lamoda_excel_file(file_path: str, data_list: List[Prod]):
    """
    Updates or creates an Excel file with Lamoda product data.
    Handles existing files and avoids duplicates based on normalized URLs.
    """
    try:
        # Create directory if it doesn't exist
        directory = os.path.dirname(file_path)
        if not os.path.exists(directory):
            os.makedirs(directory)

        # Load or create workbook
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Products"

        # Headers
        headers = ["Article", "Brand", "Title", "Price", "Old Price", "New Price", "Full URL", "Update Date"]
        if sheet.max_row == 1:
            for col_num, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col_num, value=header)

        # Create a set of existing normalized URLs and their row numbers
        existing_data = {
            normalize_url(str(sheet.cell(row=row_num, column=7).value)): row_num  # Normalize the existing URL
            for row_num in range(2, sheet.max_row + 1)
        }


        # Iterate through the new data
        for item in data_list:
            row = [
                item.article,
                item.brand,
                item.title,
                item.price,
                item.price_old,
                item.price_new,
                item.full_url,
                item.update_date,
            ]

            normalized_new_url = normalize_url(item.full_url)

            # Check if the URL already exists
            if normalized_new_url in existing_data:
                existing_row_num = existing_data[normalized_new_url]
                existing_date_str = str(sheet.cell(row=existing_row_num, column=8).value)

                # Compare update dates
                if item.update_date != existing_date_str:
                    # Update existing row
                    for col_num, value in enumerate(row, 1):
                         sheet.cell(row=existing_row_num, column=col_num, value=value)
                    log(f"Updated in Excel: {item.title} (new date: {item.update_date})")
                else:
                     log(f"Skipped (up-to-date): {item.title}")
            else:
                # Add new row
                sheet.append(row)
                log(f"Added to Excel: {item.title}")

        workbook.save(file_path)
        log(f"Excel update complete: {file_path}")
        send_debug(f"Excel lamoda успешно обновлен: {file_path}")
        return True  # Indicate success

    except Exception as e:
        log(f"Error updating Excel file: {e}")
        send_debug(f"Ошибка сохранения Excel lamoda: {e}")
        return False  # Indicate failure