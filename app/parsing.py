import time
import json
import os
from selenium import webdriver
from selenium_stealth import stealth
from bs4 import BeautifulSoup
import requests
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium import webdriver
from datetime import datetime
from pathlib import Path
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import re

import shutil
try:
    from .telegram_debugger import send_debug
except:
    from telegram_debugger import send_debug
import random


COOKIES_FILE = "ozon_cookies.json"  # Путь к файлу cookies

def init_webdriver():
    try:
        options = Options()
        # Заголовки
        options.add_argument("user-agent=Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/132.0.6834.83 Safari/537.36")
        options.add_argument("accept-language=ru,en-US;q=0.9,en;q=0.8")
        options.add_argument("accept-encoding=gzip, deflate, br, zstd")
        options.add_argument("content-type=application/json")
        options.add_argument("sec-ch-ua-platform=\"Linux\"")

        profile_dir = f"/tmp/chrome_profile_{os.getpid()}"
        if os.path.exists(profile_dir):
            shutil.rmtree(profile_dir)
        options.add_argument(f"--user-data-dir={profile_dir}")

        # Путь к локально установленному chromedriver
        chromedriver_path = "/usr/local/bin/chromedriver"  # Укажите правильный путь

        # Инициализация драйвера с параметрами
        driver = webdriver.Chrome(service=Service(chromedriver_path), options=options)

        # Применяем stealth для обхода обнаружения
        stealth(driver,
                languages=["en-US", "en"],
                vendor="Google Inc.",
                platform="Linux",
                webgl_vendor="Intel Inc.",
                renderer="Intel Iris OpenGL Engine",
                fix_hairline=True)

        driver.maximize_window()
        return driver
    except Exception as e:
        print(f"Chromedriver error: {e}")
        send_debug(f"Chromedriver error: {e}")
        return None

def init_apidriver():
    try:
        options = Options()
        # Заголовки
        options.add_argument("user-agent=Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/132.0.6834.83 Safari/537.36")
        options.add_argument("accept-language=ru,en-US;q=0.9,en;q=0.8")
        options.add_argument("accept-encoding=gzip, deflate, br, zstd")
        options.add_argument("content-type=application/json")
        options.add_argument("sec-ch-ua-platform=\"Linux\"")

        profile_dir = f"/tmp/chrome_profile_{os.getpid()}"
        if os.path.exists(profile_dir):
            shutil.rmtree(profile_dir)
        options.add_argument(f"--user-data-dir={profile_dir}")

        # Путь к локально установленному chromedriver
        chromedriver_path = "/usr/local/bin/chromedriver"  # Укажите правильный путь

        # Инициализация драйвера с параметрами
        driver = webdriver.Chrome(service=Service(chromedriver_path), options=options)

        # Применяем stealth для обхода обнаружения
        stealth(driver,
                languages=["en-US", "en"],
                vendor="Google Inc.",
                platform="Linux",
                webgl_vendor="Intel Inc.",
                renderer="Intel Iris OpenGL Engine",
                fix_hairline=True)

        driver.minimize_window()
        return driver
    except Exception as e:
        print(f"Chromeapidriver error: {e}")
        send_debug(f"Chromeapidriver error: {e}")
        return None

api = None

def change_api(new_api):
    global api
    api = new_api

def quit_api():
    global api
    api.quit()

def save_cookies(obj, path):
    """Сохраняет cookies для Selenium или Requests в файл в формате JSON."""
    try:
        if isinstance(obj, webdriver.Chrome):  # Проверка типа через конкретный драйвер
            selenium_cookies = obj.get_cookies()  # Получаем cookies из Selenium
            # Конвертируем в RequestsCookies
            requests_cookies = {cookie['name']: cookie['value'] for cookie in selenium_cookies}

            # Записываем cookies в файл в формате JSON
            with open(path, "w", encoding="utf-8") as file:
                json.dump({'selenium': selenium_cookies, 'requests': requests_cookies}, file, ensure_ascii=False, indent=4)

        elif isinstance(obj, requests.Session):  # Если это Requests Session
            requests_cookies = obj.cookies  # Получаем cookies из Requests
            selenium_cookies = [{'name': cookie.name, 'value': cookie.value, 'domain': cookie.domain, 'path': cookie.path}
                                for cookie in requests_cookies]

            # Записываем cookies в файл в формате JSON
            with open(path, "w", encoding="utf-8") as file:
                json.dump({'selenium': selenium_cookies, 'requests': requests_cookies.get_dict()}, file, ensure_ascii=False, indent=4)

        print("Cookies успешно сохранены в JSON.")

    except Exception as e:
        print(f"Ошибка при сохранении cookies: {e.__class__.__name__} — {e}")
        send_debug(f"Ошибка при сохранении cookies: {e.__class__.__name__} — {e}")



def load_cookies(obj, path):
    """Загружает cookies для Selenium или Requests из файла в формате JSON."""
    try:
        # Проверка наличия файла
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as file:
                data = json.load(file)
                selenium_cookies = data.get('selenium', [])
                requests_cookies_dict = data.get('requests', {})

                if isinstance(obj, webdriver.Chrome):  # Если это Selenium WebDriver
                    # Добавление cookies в Selenium
                    for cookie in selenium_cookies:
                        try:
                            obj.add_cookie(cookie)
                        except Exception as e:
                            print(f"Ошибка при добавлении cookies в Selenium: {e.__class__.__name__} — {e}")
                elif isinstance(obj, requests.Session):  # Если это Requests Session
                    # Добавление cookies в Requests
                    for name, value in requests_cookies_dict.items():
                        obj.cookies.set(name, value)

                print("Cookies успешно загружены из JSON.")
   
    except Exception as e:
        print(f"Ошибка при загрузке cookies: {e.__class__.__name__} — {e}")
        send_debug(f"Ошибка при загрузке cookies. file will be removed: {e.__class__.__name__} — {e}")
        if os.path.exists(path):
            os.remove(path)  # Удаляет поврежденный файл
            print(f"Удален поврежденный файл cookies: {path}")


def get_product_info(product_url):
    try:
        api.get(f"https://www.ozon.ru/api/composer-api.bx/page/json/v2?url={product_url}")
        # Ждем, пока появится элемент <pre> (до 10 секунд)
        pre_element = WebDriverWait(api, 10).until(
            EC.presence_of_element_located((By.TAG_NAME, "pre"))
        )

        # Получаем текст из <pre>
        content = pre_element.text
        
        

        # Проверяем на антибот-защиту
        if "challenge" in content.lower():
            print("Сработала защита от ботов.")
            return None

        try:
            json_data = json.loads(content)
            #with open("card.json", 'w', encoding='utf-8') as f:
            #    json.dump(json_data, f, ensure_ascii=False, indent=4)
        except json.JSONDecodeError as e:
            print(f"Ошибка при разборе JSON: {e}")
            return None
        
        product_info = {}

        try:
            #извлечение widgetstates класса
            widget_states = json_data.get("widgetStates", {})
        except Exception as e:
            print(f"Ошибка при извлечении widgetstates: {e}")

        #подъизвлечение и заполнение
       
        try:
            # Название товара

            # Шаблон для поиска ключей с дефисом во избежание повторов
            prefix = "webProductHeading-"

            
            for key, value in widget_states.items():
                if key.startswith(prefix):
                    try:
                        # Декодируем JSON-строку в словарь
                        decoded_values = json.loads(value)
                    except json.JSONDecodeError:
                        print(f"Ошибка декодирования JSON для ключа: {key}")

            # доступ к конкретным элементам
            gotval = decoded_values.get("title")

            #применение данных
            product_info['name'] = gotval
            print(f"Название товара: OK")
        
        except Exception as e:
            product_info['name'] = "Не указано"
            print(f"Название товара: Не удалось спарсить. Ошибка : {e}")
      
        try:
            # Цена со скидкой

            # Шаблон для поиска ключей с дефисом во избежание повторов
            prefix = "webPrice-"

            
            for key, value in widget_states.items():
                if key.startswith(prefix):
                    try:
                        # Декодируем JSON-строку в словарь
                        decoded_values = json.loads(value)
                    except json.JSONDecodeError:
                        print(f"Ошибка декодирования JSON для ключа: {key}")

            # доступ к конкретным элементам
            gotvalSalePrice = decoded_values.get("cardPrice")

            #применение данных
            product_info['discounted_price'] = gotvalSalePrice.strip().replace(' ', '')
            print(f"Цена со скидкой: OK")
        
        except Exception as e:
            product_info['discounted_price'] = "Не указано"
            print(f"Цена со скидкой: Не удалось спарсить. Ошибка : {e}")

        try:
            # Цена без скидки

            # Шаблон для поиска ключей с дефисом во избежание повторов
            prefix = "webPrice-"

            
            for key, value in widget_states.items():
                if key.startswith(prefix):
                    try:
                        # Декодируем JSON-строку в словарь
                        decoded_values = json.loads(value)
                    except json.JSONDecodeError:
                        print(f"Ошибка декодирования JSON для ключа: {key}")

            # доступ к конкретным элементам
            gotvalPrice = decoded_values.get("price")

            #применение данных
            product_info['original_price'] = gotvalPrice.strip().replace(' ', '')
            print(f"Цена без скидки: OK")
        
        except Exception as e:
            product_info['original_price'] = "Не указано"
            print(f"Цена без скидки: Не удалось спарсить. Ошибка : {e}") 

        try:
            # Скидка

            # Шаблон для поиска ключей с дефисом во избежание повторов
            prefix = "webMarketingLabels-"

            
            for key, value in widget_states.items():
                if key.startswith(prefix):
                    try:
                        # Декодируем JSON-строку в словарь
                        decoded_values = json.loads(value)
                    except json.JSONDecodeError:
                        print(f"Ошибка декодирования JSON для ключа: {key}")

            # доступ к конкретным элементам
            discount_text = decoded_values['labels'][0]['badge']['text']
            gotval = discount_text.replace('−', '')

            #применение данных
            product_info['discount'] = gotval
            print(f"Скидка: OK")
        
        except Exception as e:
            product_info['discount'] = "Нет данных"
            print(f"Скидка: Не удалось спарсить. Ошибка : {e}")
       
        try:
            # Наличие товара

            # Шаблон для поиска ключей с дефисом во избежание повторов
            prefix = "bigPromoPDP-"

            
            for key, value in widget_states.items():
                if key.startswith(prefix):
                    try:
                        # Декодируем JSON-строку в словарь
                        decoded_values = json.loads(value)
                    except json.JSONDecodeError:
                        print(f"Ошибка декодирования JSON для ключа: {key}")

            # доступ к конкретным элементам
            gotval = int(decoded_values['stockNumber']['text'])

            #применение данных
            product_info['stock_left'] = gotval
            print(f"Наличие товара: OK")
        
        except Exception as e:
            product_info['stock_left'] = "Нет данных"
            print(f"Наличие товара: Не удалось спарсить. Ошибка : {e}")


        try:
            # Параметры рассрочки

            # Шаблон для поиска ключей с дефисом во избежание повторов
            prefix = "webInstallmentPurchase-"

            
            for key, value in widget_states.items():
                if key.startswith(prefix):
                    try:
                        # Декодируем JSON-строку в словарь
                        decoded_values = json.loads(value)
                    except json.JSONDecodeError:
                        print(f"Ошибка декодирования JSON для ключа: {key}")

            # Найти сумму платежа
            payment_text = None
            if 'titleRs' in decoded_values:
                for item in decoded_values['titleRs']:
                    if 'content' in item and 'платеж' in item['content']:
                        payment_text = item['content']
                        break
            
            # Извлечь только сумму из текста
            if payment_text:
                import re
                amount = re.search(r'≈([\d\s]+)₽', payment_text)
                if amount:
                    payment_amount = amount.group(1).strip().replace(' ', '')
                    product_info['installment'] = f"{payment_amount} ₽"
                    print(f"Параметры рассрочки: OK")
                else:
                    product_info['installment'] = "Нет данных"
                    print("Параметры рассрочки: Не удалось спарсить. Сумма платежа не найдена.")
            else:
                product_info['installment'] = "Нет данных"
                print("Параметры рассрочки: Не удалось спарсить. Информация о платеже отсутствует.")
        
        except Exception as e:
            product_info['installment'] = "Нет данных"
            print(f"Параметры рассрочки: Не удалось спарсить. Ошибка : {e}")
       

        
        try:
            # Категория товара
            try:
                #извлечение layoutTrackingInfo класса
                layoutTrackingInfo = json.loads(json_data.get("layoutTrackingInfo", {}))
            except Exception as e:
                print(f"Ошибка при извлечении layoutTrackingInfo: {e}")
        
            # доступ к конкретным элементам
            gotval = layoutTrackingInfo.get("categoryName")

            #применение данных
            product_info['category'] = gotval
            print(f"Категория товара: OK")
        
        except Exception as e:
            product_info['category'] = "Не указано"
            print(f"Категория товара: Не удалось спарсить. Ошибка : {e}")
        

        try:
            # Производитель или бренд

            # Извлечение скрипта с innerHTML
            script = json_data.get("seo", {}).get("script", [])
            # Распарсить JSON из innerHTML
            product_data = json.loads(script[0].get("innerHTML", "{}"))
            # Извлечь бренд
            brand = product_data.get("brand")
                    

            #применение данных
            product_info['brand'] = brand
            print(f"Производитель или бренд: OK")
        
        except Exception as e:
            product_info['brand'] = "Нет данных"
            print(f"Производитель или бренд: Не удалось спарсить. Ошибка : {e}")

        try:
            # Описание

            # Извлечение скрипта с innerHTML
            script = json_data.get("seo", {}).get("script", [])
            # Распарсить JSON из innerHTML
            product_data = json.loads(script[0].get("innerHTML", "{}"))
            # Извлечь
            gotval = product_data.get("description")
                    

            #применение данных
            product_info['description'] = gotval
            print(f"Описание: OK")
        
        except Exception as e:
            product_info['description'] = "Нет данных"
            print(f"Описание: Не удалось спарсить. Ошибка : {e}")

        try:
            # Оценка

            # Шаблон для поиска ключей с дефисом во избежание повторов
            prefix = "webReviewProductScore-"

            
            for key, value in widget_states.items():
                if key.startswith(prefix):
                    try:
                        # Декодируем JSON-строку в словарь
                        decoded_values = json.loads(value)
                    except json.JSONDecodeError:
                        print(f"Ошибка декодирования JSON для ключа: {key}")


            # доступ к конкретным элементам
            total_score = decoded_values.get("totalScore")
            #применение данных
            product_info['rating'] = total_score
            print(f"Оценка: OK")
        
        except Exception as e:
            product_info['rating'] = "Нет данных"
            print(f"Оценка: Не удалось спарсить. Ошибка : {e}")

        try:
            # Отзывы

            # Шаблон для поиска ключей с дефисом во избежание повторов
            prefix = "webReviewProductScore-"

            
            for key, value in widget_states.items():
                if key.startswith(prefix):
                    try:
                        # Декодируем JSON-строку в словарь
                        decoded_values = json.loads(value)
                    except json.JSONDecodeError:
                        print(f"Ошибка декодирования JSON для ключа: {key}")


            # доступ к конкретным элементам
            reviews_count = decoded_values.get("reviewsCount")
            #применение данных
            product_info['reviews_count'] = reviews_count
            print(f"Отзывы: OK")
        
        except Exception as e:
            product_info['reviews_count'] = "Нет данных"
            print(f"Отзывы: Не удалось спарсить. Ошибка : {e}")

        try:
            # Доступные медиафайлы

            # Шаблон для поиска ключей с дефисом во избежание повторов
            prefix = "webStickyProducts-"

            
            for key, value in widget_states.items():
                if key.startswith(prefix):
                    try:
                        # Декодируем JSON-строку в словарь
                        decoded_values = json.loads(value)
                    except json.JSONDecodeError:
                        print(f"Ошибка декодирования JSON для ключа: {key}")

            # доступ к конкретным элементам
            gotval = decoded_values.get("coverImageUrl")

            #применение данных
            product_info['media_files'] = gotval
            print(f"Картинка: OK")
        
        except Exception as e:
            product_info['media_files'] = "Не указано"
            print(f"Картинка: Не удалось спарсить. Ошибка : {e}")

        print("Попытка парсинга карточки завершена.")
        #with open('product.json', 'w', encoding='utf-8') as f:
        #    json.dump(product_info, f, ensure_ascii=False, indent=4)
        return product_info

    except Exception as e:
        print(f"Ошибка при получении информации о продукте: {e}")
        return None

def scrolldown(driver, all_cards, processed_urls):
    save_cookies(driver,COOKIES_FILE)
    last_height = driver.execute_script("return document.body.scrollHeight")
    
    while True:
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        delay = random.uniform(1.9, 2.5)
        time.sleep(delay)  # Даём время для подгрузки контента
        new_height = driver.execute_script("return document.body.scrollHeight")
       
        # Парсим карточки после каждой прокрутки
        search_page_html = BeautifulSoup(driver.page_source, "html.parser")
        content = search_page_html.find("div", {"id": "layoutPage"})
        if content:
            content = content.find("div")
            
            # Находим все контейнеры с карточками
            all_containers = content.find_all("div", {"data-index": re.compile(r"^\d+$")})

            for container in all_containers:
                card_link = container.find("a", class_=re.compile(r"tile-clickable-element"), href=True)
                if card_link:
                    card_url = card_link["href"]
                    product_url = "https://ozon.ru" + card_url

                    if card_url not in processed_urls:
                        processed_urls.add(card_url)

                        # Название (если доступно на превью)
                        card_name_span = container.find("span", {"class": "tsBody500Medium"})
                        card_name = card_name_span.text.strip() if card_name_span else "No Name"

                        # Получаем полную информацию о товаре
                        try:
                            product_info = get_product_info(card_url)
                        except Exception as e:
                            print(f"Ошибка при получении данных для {card_url}: {e}")
                            product_info = {}

                        if product_info:
                            # Извлечение данных из product_info с проверками
                            name = product_info.get("name", card_name)
                            description = product_info.get("description", "Нет описания")
                            price = product_info.get("discounted_price", "Нет цены")
                            original_price = product_info.get("original_price", "Нет данных")
                            discount = product_info.get("discount", "Нет скидки")
                            stock_left = product_info.get("stock_left", "Нет данных")
                            installment = product_info.get("installment", "Нет данных")
                            category = product_info.get("category", "Нет данных")
                            brand = product_info.get("brand", "Нет данных")
                            rating = product_info.get("rating", "Нет данных")
                            rating_counter = product_info.get("reviews_count", 0)
                            image_url = product_info.get("media_files")
                            update = datetime.now().strftime("%Y-%m-%d")

                            card_info = {
                                "url": product_url,
                                "name": name,
                                "description": description,
                                "price": price,
                                "original_price": original_price,
                                "discount": discount,
                                "stock_left": stock_left,
                                "installment": installment,
                                "category": category,
                                "brand": brand,
                                "rating": rating,
                                "rating_counter": rating_counter,
                                "image_url": image_url,
                                "update": update
                            }

                            all_cards.append(card_info)
                            print(f"{name} - DONE")
                            print("Спарсено карточек:", len(all_cards))
                        else:
                            print(f"Не удалось получить данные для товара {card_url}")
                else:
                    print("Не удалось найти ссылку на карточку")
                    
        if new_height == last_height:
            break
        last_height = new_height

def get_searchpage_cards(driver, url, all_cards=None, processed_urls=None):
    if all_cards is None:
        all_cards = []  # Новый список для карточек
    if processed_urls is None:
        processed_urls = set()  # Новый набор для обработанных URL
    driver.get(url)
    time.sleep(1.2)
    try:
        # Попытка загрузить cookies
        load_cookies(driver,COOKIES_FILE)

        # Перезагрузка страницы для применения cookies
        driver.refresh()
    except Exception as e:
        print(f"Error: {e}")
    
    # Задержка от 2 до 5 секунд
    delay = random.uniform(0.5, 1)
    time.sleep(delay)
    scrolldown(driver, all_cards, processed_urls)
    
    content = BeautifulSoup(driver.page_source, "html.parser").find("div", {"id": "layoutPage"}).find("div")

    content_with_next = [div for div in content.find_all("a", href=True) if "Дальше" in str(div)]
    if not content_with_next:
        return all_cards
    else:
        next_page_url = "https://www.ozon.ru" + content_with_next[0]["href"]
        return get_searchpage_cards(driver, next_page_url, all_cards, processed_urls)
    
def parse_category(category_url):
    driver = init_webdriver()
    global api
    api = init_apidriver()
    search_cards = []
    try:
        search_cards = get_searchpage_cards(driver, category_url)
        if len(search_cards)!=0:
            print("Я успешно нашёл", len(search_cards), "товаров по ссылке " + category_url)
            send_debug(f"Парсинг {category_url} успешно завершен с {len(search_cards)} карточек")
        else:
            print("Ошибка. Я не нашёл ни одного товара по ссылке " + category_url)
            send_debug("Ошибка. Я не нашёл ни одного товара по ссылке " + category_url)
            return None
    except Exception as e:
        print(f"Произошла ошибка при парсинге: {e.__class__.__name__} — {e}")
        send_debug(f"Произошла ошибка при парсинге ozon {category_url}: {e.__class__.__name__} — {e}")

    save_cookies(driver,COOKIES_FILE)
    api.quit()
    driver.quit()

    # Преобразование данных в итоговую структуру
    parsed_data = []
    for product in search_cards:
        parsed_data.append({
            'description': product.get('description', 'Нет описания'),
            'name': product.get('name', 'Не указано'),
            'discounted_price': product.get('price', 'Нет цены'),
            'original_price': product.get('original_price', 'Нет цены'),
            'discount': product.get('discount', 'Нет скидки'),
            'stock_left': product.get('stock_left', 'Нет данных'),
            'installment': product.get('installment', 'Нет данных'),
            'category': product.get('category', 'Нет данных'),
            'brand': product.get('brand', 'Нет данных'),
            'rating': product.get('rating', 'Нет данных'),
            'reviews_count': product.get('rating_counter', 'Нет отзывов'),
            'media_file': product.get('image_url', '/static/images/no-image.png'),
            'url': product.get('url', ''),  # Добавляем URL товара
            'update': product.get('update', 'Не известна дата обновления')
        })

    CATEGORY_EXCEL_FILE_PATH = f'parsed_data/ozon/{category_url[:100].replace('/', '_').replace(':', '_')}_products.xlsx'
    #CATEGORY_FILE_PATH = f'parsed_data/ozon/{category_url.replace('/', '_').replace(':', '_')}_products.json'
    #with open(CATEGORY_FILE_PATH, 'w', encoding='utf-8') as f:
    #    json.dump(parsed_data, f, ensure_ascii=False, indent=4)

    # GSHEETS
    # Открываем Google Таблицу по её ID
    #SPREADSHEET_ID = "YOUR_GOOGLE_SPREADSHEET_ID"
    #sheet_name = f'{category_url.replace('https://www.ozon.ru/category/','').replace('/', '_').replace(':', '_')}' # Название листа

    #update_google_sheet(sheet_name, parsed_data,SPREADSHEET_ID)
    
    update_excel_file(CATEGORY_EXCEL_FILE_PATH, parsed_data)
    return parsed_data[:100] # Возвращаем первые 100 записей во избежание перенаполнения страницы


#excel

import openpyxl
from datetime import datetime


def clean_url(url):
    """Обрезает URL, оставляя только основную часть (до ?)."""
    return url.split('?')[0] if url else None  # Handle potential None values


def update_excel_file(file_path, data_list):
    """
    Обновляет Excel-файл данными из списка словарей.

    Args:
        file_path (str): Путь к Excel-файлу.
        data_list (list): Список словарей с данными о товарах.
    """
    try:
        # Загрузка или создание книги Excel
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Products"

        # Заголовки столбцов (если лист пустой)
        headers = [
            "Название", "Описание", "Цена со скидкой", "Обычная цена", "Скидка", "Остаток",
            "Рассрочка", "Категория", "Бренд", "Рейтинг", "Отзывы",
            "Обновление", "Ссылка", "Ссылка на изображение"  # Добавили "Описание"
        ]
        if sheet.max_row == 1:  # Проверка на пустой лист (только заголовки)
            for col_num, header in enumerate(headers, start=1):
                sheet.cell(row=1, column=col_num, value=header)

        # Создание словаря для быстрого поиска существующих записей по URL
        #  {url: (row_index, update_date)}
        existing_data = {}
        for row_index in range(2, sheet.max_row + 1):
            url_cell = sheet.cell(row=row_index, column=13).value  # 13 - колонка "Ссылка"
            if url_cell:
                cleaned_url = clean_url(url_cell)
                update_date_cell = sheet.cell(row=row_index, column=12).value # 12 колонка "Обновление"
                existing_data[cleaned_url] = (row_index, update_date_cell)


        # Перебор данных для добавления/обновления
        for item in data_list:
            cleaned_url = clean_url(item['url'])  # Очищенный URL
            if not cleaned_url:
                print(f"Пропущено: {item['name']} (отсутствует URL)")
                continue

            row_data = [
                item["name"],
                item["description"],
                item["discounted_price"],
                item["original_price"],
                item["discount"],
                item["stock_left"],
                item["installment"],
                item["category"],
                item["brand"],
                item["rating"],
                item["reviews_count"],
                item["update"],
                item["url"],
                item["media_file"]
                # Добавили значение для столбца "Описание"
            ]

            # Проверка наличия записи в Excel
            if cleaned_url in existing_data:
                row_index, existing_update_date = existing_data[cleaned_url]

                # Сравнение дат обновления
                if item["update"] != existing_update_date:
                    # Обновление существующей строки
                    for col_num, value in enumerate(row_data, start=1):
                        sheet.cell(row=row_index, column=col_num, value=value)
                    print(f"Обновлено в Excel: {item['name']} (новая дата: {item['update']})")
                else:
                    print(f"Пропущено: {item['name']} (актуальная дата совпадает)")
            else:
                # Добавление новой строки
                sheet.append(row_data)
                print(f"Добавлено в Excel: {item['name']}")

        workbook.save(file_path)
        print(f"Обновление в {file_path} завершено!")
        #send_debug(f"Сохранение ozon категории в файл {file_path} было успешно завершено") #убрал вызов непонятной функции


    except Exception as e:
        print(f"Ошибка при обновлении {file_path}: {e}")
        #send_debug(f"Ошибка при обновлении Excel - ozon {file_path}: {e}") #убрал вызов непонятной функции




'''ЗАПИСЬ 
  В ГУГЛ  
  ТАБЛИЦЫ

import gspread
from google.oauth2.service_account import Credentials
from datetime import datetime
import re  # Для очистки даты

# Укажи путь к JSON-файлу с учетными данными
SERVICE_ACCOUNT_FILE = "credentials.json"
SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]

# Авторизация
creds = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
client = gspread.authorize(creds)


def update_google_sheet(sheet_name, data_list, spreadsheet_id):
    spreadsheet = client.open_by_key(spreadsheet_id)
    try:
        # Проверяем, есть ли лист с таким названием, если нет — создаем
        try:
            sheet = spreadsheet.worksheet(sheet_name)
        except gspread.exceptions.WorksheetNotFound:
            sheet = spreadsheet.add_worksheet(title=sheet_name, rows="1000", cols="15")

        # Заголовки (если нужно)
        headers = ["Название", "Цена со скидкой", "Обычная цена", "Скидка", "Остаток", "Рассрочка", 
                   "Категория", "Бренд", "Рейтинг", "Отзывы", "Обновление", "Ссылка", "Ссылка на изображение"]
        if not sheet.row_values(1):  # Если первая строка пустая — записываем заголовки
            sheet.append_row(headers)

        existing_data = sheet.get_all_values()  # Получаем существующие данные
        
        # Создаем словарь {url: (номер строки, дата обновления)}
        existing_urls = {clean_url(row[11].strip()): (index + 1, row[10]) for index, row in enumerate(existing_data[1:], start=2)}
        for item in data_list:
            row = [
                item["name"], item["discounted_price"], item["original_price"],
                item["discount"], item["stock_left"], item["installment"],
                item["category"], item["brand"], item["rating"],
                item["reviews_count"], item["update"], item["url"], item["media_file"]
            ]
            
            new_url = clean_url(item["url"].strip())  # Очищаем URL от параметров
            new_date_str = item["update"]

            if not new_url:
                print(f"Пропущено: {item['name']} (отсутствует URL)")
                continue

            # Флаг, указывающий, что обновление или добавление необходимо
            update_needed = False
            add_needed = True
            row_index_to_update = None

            for existing_url, (row_index, existing_date_str) in existing_urls.items():
                if existing_url == new_url:  # Нашли совпадение по основному URL
                    # Сравниваем дату обновления
                    if new_date_str != existing_date_str:
                        update_needed = True
                        row_index_to_update = row_index
                    else:
                        print(f"Пропущено: {item['name']} (актуальная дата GSHEETS совпадает)")
                    break  # Если URL найден, больше не проверяем
            
            for existing_url, (row_index, existing_date_str) in existing_urls.items():
                if existing_url == new_url:  # Нашли совпадение по основному URL
                    add_needed =False
                    break  # Если URL найден, больше не проверяем

            # Если данные отличаются или строки нет в таблице — обновляем или добавляем строку
            if update_needed:
                sheet.update(f"A{row_index_to_update}:L{row_index_to_update}", [row])
                print(f"Обновлено в GSHEETS: {item['name']} (новая дата: {new_date_str})")
            elif add_needed:
                # Если товара нет в таблице — добавляем новую строку
                sheet.append_row(row)
                print(f"Добавлено в GSHEETS: {item['name']}")

        print(f"Обновление {sheet_name} завершено!")

    except Exception as e:
        print(f"Ошибка при обновлении {sheet_name}: {e}")'''
