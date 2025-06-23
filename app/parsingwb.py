import time
import os
import requests
from datetime import datetime
from retry import retry
try:
    from .parsing import save_cookies, load_cookies
except: from parsing import save_cookies, load_cookies
import random
import openpyxl  # Добавили openpyxl
from openpyxl.utils import get_column_letter  # Добавили для работы с Excel
try:
    from .telegram_debugger import send_debug
except: from telegram_debugger import send_debug
from urllib.parse import urlparse, parse_qs, urlencode
COOKIES_FILE = "wb_cookies.json"  # Путь к файлу cookies

session = requests.Session()

def get_catalogs_wb() -> dict:
    """получаем полный каталог Wildberries"""
    url = 'https://static-basket-01.wbbasket.ru/vol0/data/main-menu-ru-ru-v3.json'
    headers = {
        "Accept": "application/json",
        "Accept-Encoding": "gzip, deflate, br, zstd",
        "Accept-Language": "ru,en-US;q=0.9,en;q=0.8",
        "Sec-CH-UA": '"Chromium";v="132", "Not_A Brand";v="24"',
        "Sec-CH-UA-Platform": '"Linux"',
        "User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/132.0.6834.83 Safari/537.36"
    }

    load_cookies(session,  COOKIES_FILE)
    try:
        requested = session.get(url, headers=headers).json()
        save_cookies(session, COOKIES_FILE)
    except Exception as e:
        print("Error during requesting catalog:", e)
        send_debug("Ошибка получения вб каталога:", e)
        requested = None
    return requested


def get_data_category(catalogs_wb: dict) -> list:
    """сбор данных категорий из каталога Wildberries"""
    catalog_data = []
    if isinstance(catalogs_wb, dict) and 'childs' not in catalogs_wb:
        catalog_data.append({
            'name': f"{catalogs_wb['name']}",
            'shard': catalogs_wb.get('shard', None),
            'url': catalogs_wb['url'],
            'query': catalogs_wb.get('query', None)
        })
    elif isinstance(catalogs_wb, dict):
        catalog_data.append({
            'name': f"{catalogs_wb['name']}",
            'shard': catalogs_wb.get('shard', None),
            'url': catalogs_wb['url'],
            'query': catalogs_wb.get('query', None)
        })
        catalog_data.extend(get_data_category(catalogs_wb['childs']))
    else:
        for child in catalogs_wb:
            catalog_data.extend(get_data_category(child))
    return catalog_data


def search_category_in_catalog(url: str, catalog_list: list) -> dict:
    """проверка пользовательской ссылки на наличии в каталоге"""
    for catalog in catalog_list:
            if catalog['url'] == url.split('https://www.wildberries.ru')[-1]:
                print(f'найдено совпадение: {catalog["name"]}')
                return catalog
    print("Ошибка: в каталоге не найдено совпадений")
    send_debug(f"Ошибка: в каталоге wb не найдено совпадений {url}")
    return None  # Возвращаем None, если категория не найдена


def get_data_from_json(json_file: dict) -> list:
    """извлекаем из json данные"""
    try:
        data_list = []
        for data in json_file['data']['products']:
            sku = data.get('id')
            name = data.get('name')
            price = int(data.get("priceU", 0) / 100)  # Добавил обработку отсутствия ключа
            salePriceU = int(data.get('salePriceU', 0) / 100)  # Добавил обработку отсутствия ключа
            cashback = data.get('feedbackPoints')
            sale = data.get('sale')
            brand = data.get('brand')
            rating = data.get('rating')
            supplier = data.get('supplier')
            supplierRating = data.get('supplierRating')
            feedbacks = data.get('feedbacks')
            reviewRating = data.get('reviewRating')
            promoTextCard = data.get('promoTextCard')
            promoTextCat = data.get('promoTextCat')
            update = f"{datetime.now().strftime("%Y-%m-%d")}"
            data_list.append({
                'id': sku,
                'name': name,
                'price': price,
                'salePriceU': salePriceU,
                'cashback': cashback,
                'sale': sale,
                'brand': brand,
                'rating': rating,
                'supplier': supplier,
                'supplierRating': supplierRating,
                'feedbacks': feedbacks,
                'reviewRating': reviewRating,
                'promoTextCard': promoTextCard,
                'promoTextCat': promoTextCat,
                'update': update,
                'link': f'https://www.wildberries.ru/catalog/{data.get("id")}/detail.aspx?targetUrl=BP'
            })
    except Exception as e:
        print(f"Ошибка парсинга полученных из вб данных {json_file}")
        send_debug(f"Ошибка парсинга полученных из вб данных {json_file}")
    return data_list

HTTPANSWER = 0

@retry(Exception, tries=-1, delay=0)
def scrap_page(page: int, shard: str, query: str,paramsURL: str) -> dict:
    """Сбор данных со страниц"""
    global HTTPANSWER
    headers = {
        "Accept": "application/json",
        "Accept-Encoding": "gzip, deflate, br, zstd",
        "Accept-Language": "ru,en-US;q=0.9,en;q=0.8",
        "Sec-CH-UA": '"Chromium";v="132", "Not_A Brand";v="24"',
        "Sec-CH-UA-Platform": '"Linux"',
        "User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/132.0.6834.83 Safari/537.36"
    }
    url = f'https://catalog.wb.ru/catalog/{shard}/catalog?page={page}&{paramsURL}&{query}'
    load_cookies(session, COOKIES_FILE)
    try:
        r = session.get(url, headers=headers)
        save_cookies(session, COOKIES_FILE)
    except Exception as e:
        print(f'Error on requesting page {page} {e}')
        send_debug(f'Ошибка получения страницы вб {page} {e}')
    
    print(f'Статус: {r.status_code} Страница {page} Идет сбор...')
    # Добавляем проверку на успешный статус ответа
    HTTPANSWER = r.status_code
    if HTTPANSWER == 200:
        return r.json()
    elif HTTPANSWER == 429:
        print(f'Попытка парсинга страницы {page} заблокирована')
        return {'data': {'products': []}} # Возвращаем пустой список
    else:
        print(f'Попытка парсинга страницы {page} не удалась')
        return {'data': {'products': []}} # Возвращаем пустой список, если запрос не успешен
    


def parse_category_wb(category_url: str) -> dict:
    """Основная функция для парсинга категории WB."""
    global HTTPANSWER
    # Разбираем URL и параметры
    parsed_url = urlparse(category_url)
    source_params = parse_qs(parsed_url.query)

    # Исходные параметры из твоего кода
    base_params = {
        "appType": "1",
        "curr": "rub",
        "dest": "-1257786",
        "locale": "ru",
        "sort": "popular",
        "spp": "0",
    }

    # Объединяем параметры, НЕ перезаписывая существующие
    for key, value in source_params.items():
        if key not in base_params:
            base_params[key] = value[0] if len(value) == 1 else value

    # Создаём строку параметров
    paramsURL = urlencode(base_params, doseq=True)

    if 'digital.wildberries.ru' in category_url:
        from .parsingwb_digital import parse_wildberries_category
        data_list = parse_wildberries_category(category_url,COOKIES_FILE,session)
    else:
        catalog_data = get_data_category(get_catalogs_wb())
        category = search_category_in_catalog(url=clean_url(category_url), catalog_list=catalog_data)

        if not category:
            print('Ошибка! Неверно указана ссылка на категорию.')
            send_debug(f'Ошибка wb! Неверно указана ссылка на категорию {category_url}.')
            return {}

        page = 1
        data_list = []

        retry_count = 0  # Счётчик неудачных попыток
        max_delay = 600  # 10 минут в секундах
        base_delay = 1.2   # Начальная задержка
        max_wait_reached_count = 0  # Счётчик случаев, когда достигли максимальной задержки

        while True:
            data = scrap_page(page=page, shard=category['shard'], query=category['query'], paramsURL=paramsURL)

            if data and data['data']['products']:
                new_items = get_data_from_json(data)
                print(f'Добавлено позиций: {len(new_items)} с {page}-й страницы')
                data_list.extend(new_items)
                page += 1  # Переход к следующей странице
                retry_count = 0  # Сбрасываем счётчик неудач при успешном запросе
                max_wait_reached_count = 0  # Сбрасываем счётчик 10-минутных ожиданий

            elif HTTPANSWER == 429:
                retry_count += 1  # Увеличиваем счётчик при каждой 429 ошибке
                delay = min(base_delay * 2**retry_count, max_delay)  # Экспоненциальный рост, но не больше 10 минут
                delay += random.uniform(-delay * 0.2, delay * 0.2)  # Добавляем небольшую случайность (±20%)

                print(f'Страница {page} вернула {HTTPANSWER}. Ждём {int(delay)} сек. Попытка #{retry_count}')
                time.sleep(delay)

                if delay >= max_delay:
                    max_wait_reached_count += 1
                    if max_wait_reached_count >= 2:
                        print(f'Достигнуто два 10-минутных ожидания без результатов. Остановка парсинга.')
                        break

            else:
                print(f'Страница {page} всё ещё пуста. Остановка парсинга.')
                break  # Завершаем цикл



    
    if len(data_list)!=0:
        print(f'Сбор данных завершен. Собрано: {len(data_list)} товаров.')
        send_debug(f'Сбор данных wb {category_url} завершен. Собрано: {len(data_list)} товаров.')
    else:
        print('Сбор данных wb не удался.')
        send_debug('Сбор данных wb не удался.')
        return None
    #Сохранение в Excel
    CATEGORY_EXCEL_FILE_PATH = f'parsed_data/wb/{category_url[:100].replace("/", "_").replace(":", "_")}_products.xlsx'
    update_excel_file(CATEGORY_EXCEL_FILE_PATH, data_list)

    return data_list[:100] # Возвращаем первые 100 записей во избежание перенаполнения страницы

def clean_url(url):
    """Обрезает URL, оставляя только основную часть (до ?).  Для WB это не нужно, но функция пусть будет."""
    return url.split('?')[0]

def update_excel_file(file_path, data_list):
    """Обновление Excel файла данными."""
    try:
        # Создаем папку, если она не существует
        directory = os.path.dirname(file_path)
        if not os.path.exists(directory):
            os.makedirs(directory)

        # Проверяем, существует ли файл, если нет, создаем новый
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Products"

        # Заголовки
        headers = [
            'id', 'Название', 'Цена', 'Цена со скидкой', 'Кэшбек', 'Скидка',
            'Бренд', 'Рейтинг', 'Продавец', 'Рейтинг продавца', 'Отзывы',
            'Рейтинг по отзывам', 'Промо (Карточка)', 'Промо (Категория)', 'Обновление', 'Ссылка'
        ]

        if sheet.max_row == 1:  # Если первый ряд пустой, записываем заголовки
            for col_num, header in enumerate(headers, start=1):
                sheet[get_column_letter(col_num) + "1"] = header

        # Получаем существующие данные (ссылки и даты обновления)
        existing_data = {}
        for index, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=len(headers)), start=2):
            existing_url = row[15].value  #  Предполагаем, что ссылка в 16 столбце (индекс 15)
            if existing_url:  # проверка на None
               existing_url = clean_url(existing_url.strip())
               existing_data[existing_url] = (index, row[14].value)  # Дата в 15 столбце

        for item in data_list:
            row = [
                item['id'], item['name'], item['price'], item['salePriceU'],
                item['cashback'], item['sale'], item['brand'], item['rating'],
                item['supplier'], item['supplierRating'], item['feedbacks'],
                item['reviewRating'], item['promoTextCard'], item['promoTextCat'],
                item['update'], item['link']
            ]

            new_url = clean_url(item['link'].strip())  # Ссылка на товар
            new_date_str = item['update']  # Дата обновления

            if not new_url:
                print(f"Пропущено: {item['name']} (отсутствует URL)")
                continue

            # Флаги для определения, нужно ли обновление или добавление
            update_needed = False
            add_needed = True

            # Индекс строки для обновления (если нужно)
            row_index_to_update = None

            for existing_url, (row_index, existing_date_str) in existing_data.items():
                if existing_url == new_url:
                    add_needed = False  # Нашли совпадение - добавлять не нужно
                    # Сравниваем дату обновления (если она строковая)
                    if isinstance(existing_date_str, str) and new_date_str != existing_date_str:
                         update_needed = True
                         row_index_to_update = row_index  # Запоминаем индекс
                    elif isinstance(existing_date_str, datetime) and new_date_str != existing_date_str.strftime("%Y-%m-%d"):
                        # Если existing_date_str это datetime, сравниваем корректно
                        update_needed = True
                        row_index_to_update = row_index

                    break

            if update_needed:
                # Обновляем строку
                for col_num, value in enumerate(row, start=1):
                    sheet[get_column_letter(col_num) + str(row_index_to_update)] = value
                print(f"Обновлено в Excel: {item['name']} (новая дата: {new_date_str})")

            elif add_needed:
                # Добавляем новую строку
                sheet.append(row)
                print(f"Добавлено в Excel: {item['name']}")


        workbook.save(file_path)  # Сохраняем файл
        print(f"Обновление в {file_path} завершено!")
        send_debug(f"Обновление wb excel в {file_path} завершено!")


    except Exception as e:
        print(f"Ошибка при обновлении {file_path}: {e}")
        send_debug(f"Ошибка при обновлении wb excel {file_path}: {e}")